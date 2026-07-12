"""Exports partagés FDSU — aucun export fictif.

Formats réellement générés :
- Excel (.xlsx) pour dossier de décision
PDF / PowerPoint : capacité absente → boutons désactivés côté UI (pas de faux succès).
"""

from __future__ import annotations

import io
import logging
import re
from datetime import datetime, timezone
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font

logger = logging.getLogger(__name__)

ENGINE_VERSION = "exports-1.0.0"

CAPABILITIES = {
    "export_excel": True,
    "export_pdf": False,
    "export_powerpoint": False,
    "mission_planning": False,
    "simulation": False,
    "comparison": False,
    "map_navigation": True,
}


def _now() -> str:
    return datetime.now(timezone.utc).isoformat()


def get_capabilities() -> dict[str, Any]:
    return {
        "_meta": {"version": ENGINE_VERSION, "updated_at": _now()},
        "capabilities": dict(CAPABILITIES),
        "notes": {
            "export_pdf": "Export PDF non encore activé pour les dossiers",
            "export_powerpoint": "Export PowerPoint non encore activé pour les dossiers",
            "mission_planning": "Workflow mission non branché — action masquée",
            "simulation": "Simulation non branchée — action masquée",
        },
    }


def _safe_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, dict):
        for key in ("text", "recommendation", "label", "title", "detail", "display", "value"):
            if value.get(key) not in (None, ""):
                return _safe_text(value.get(key))
        return ""
    if isinstance(value, (list, tuple)):
        parts = [_safe_text(v) for v in value]
        return " · ".join(p for p in parts if p)
    return str(value)


def _slug(value: str) -> str:
    text = re.sub(r"[^\w\-]+", "_", str(value or "").strip(), flags=re.UNICODE)
    return re.sub(r"_+", "_", text).strip("_") or "dossier"


def build_decision_case_excel(
    entity_type: str,
    entity_id: str,
    *,
    program_code: str | None = None,
) -> tuple[bytes, str, dict[str, Any]]:
    """Génère un .xlsx réel pour le dossier de décision courant."""
    from api.services import explainable_decision_service as eds

    case = eds.get_decision_case(entity_id, asset_type=entity_type, program_code=program_code)
    if not case:
        raise LookupError("Dossier de décision introuvable.")

    asset = case.get("asset") or case.get("site") or {}
    summary = case.get("summary") or {}
    score = case.get("score") or {}
    confidence = case.get("confidence") or {}
    justification = case.get("justification") or case.get("criteria") or []
    if isinstance(justification, dict):
        justification = list(justification.values())

    wb = Workbook()
    ws = wb.active
    ws.title = "Dossier"

    header_font = Font(bold=True)
    rows = [
        ("Champ", "Valeur"),
        ("Identité", _safe_text(asset.get("site_name") or asset.get("name") or entity_id)),
        ("Code site", _safe_text(asset.get("site_code") or asset.get("business_id") or entity_id)),
        ("Type d’entité", entity_type),
        ("Programme", _safe_text(asset.get("program_code") or program_code)),
        ("Province", _safe_text(asset.get("province"))),
        ("Territoire", _safe_text(asset.get("territoire"))),
        ("Score", _safe_text(score.get("global") if isinstance(score, dict) else score)),
        ("Priorité", _safe_text(score.get("priority_label") or score.get("priority_level") or summary.get("priority_label"))),
        ("Confiance", _safe_text(confidence.get("level") if isinstance(confidence, dict) else confidence)),
        ("Recommandation", _safe_text(summary.get("recommendation") or summary.get("text") or case.get("recommendation_text"))),
        ("Besoins / impacts", _safe_text(case.get("impacts"))),
        ("Lacunes", _safe_text((confidence.get("missing") if isinstance(confidence, dict) else None) or case.get("missing_data") or "Aucune lacune identifiée")),
        ("Sources", _safe_text(case.get("sources"))),
        ("Matrice", _safe_text((case.get("matrix") or {}).get("id") or "Matrice de priorisation des sites FDSU")),
        ("Date de génération", _safe_text(case.get("generated_at") or _now())),
        ("Export généré le", _now()),
        ("Moteur", _safe_text(case.get("engine_version") or ENGINE_VERSION)),
    ]
    for r_idx, row in enumerate(rows, start=1):
        for c_idx, cell in enumerate(row, start=1):
            ws.cell(r_idx, c_idx, cell)
            if r_idx == 1:
                ws.cell(r_idx, c_idx).font = header_font

    ws2 = wb.create_sheet("Critères")
    ws2.append(["Critère", "Contribution", "Justification"])
    for cell in ws2[1]:
        cell.font = header_font
    for item in justification:
        if not isinstance(item, dict):
            ws2.append([_safe_text(item), "", ""])
            continue
        ws2.append([
            _safe_text(item.get("label") or item.get("criterion_id")),
            _safe_text(item.get("contribution_display") or item.get("contribution") or item.get("score_display")),
            _safe_text(item.get("why") or item.get("description")),
        ])

    ws3 = wb.create_sheet("Risques")
    ws3.append(["Risque"])
    ws3["A1"].font = header_font
    risks = case.get("risks") or []
    if not risks:
        ws3.append(["Aucun risque critique signalé"])
    else:
        for risk in risks:
            ws3.append([_safe_text(risk.get("label") if isinstance(risk, dict) else risk)])

    buffer = io.BytesIO()
    wb.save(buffer)
    payload = buffer.getvalue()

    site_label = _safe_text(asset.get("site_code") or asset.get("site_name") or entity_id)
    program = _safe_text(asset.get("program_code") or program_code or "fdsu")
    stamp = datetime.now().strftime("%Y%m%d")
    filename = f"Dossier_decision_{_slug(program)}_{_slug(site_label)}_{stamp}.xlsx"

    meta = {
        "bytes": len(payload),
        "filename": filename,
        "case_id": case.get("case_id"),
        "entity_type": entity_type,
        "entity_id": entity_id,
        "generated_at": _now(),
        "format": "xlsx",
    }
    logger.info(
        "export.decision_case.excel case_id=%s entity=%s/%s bytes=%s",
        case.get("case_id"),
        entity_type,
        entity_id,
        len(payload),
    )
    return payload, filename, meta
