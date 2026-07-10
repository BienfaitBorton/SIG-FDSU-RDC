"""Géocodage Intelligent FDSU — analyse Excel, contrôle qualité, correction contrôlée."""

from __future__ import annotations

import json
import re
import time
import uuid
import urllib.parse
import urllib.request
from dataclasses import asdict, dataclass, field
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.styles import Font

from api.config import DATA_MODE, connect_db

PROJECT_ROOT = Path(__file__).resolve().parents[2]
IMPORT_DIR = PROJECT_ROOT / "data" / "imports" / "geocoding"
EXPORT_DIR = PROJECT_ROOT / "data" / "exports" / "geocoding"
JOBS_DIR = EXPORT_DIR / "jobs"

# Emprise approximative RDC (WGS84)
RDC_LAT_MIN, RDC_LAT_MAX = -13.6, 5.6
RDC_LON_MIN, RDC_LON_MAX = 12.0, 31.8

LAT_ALIASES = ("latitude", "lat", "y", "coord_y", "coordy", "lat_dd", "latitud")
LON_ALIASES = ("longitude", "lon", "lng", "long", "x", "coord_x", "coordx", "lon_dd", "longitud")
ADDRESS_ALIASES = {
    "adresse": ("adresse", "address", "addr", "voie", "avenue", "av"),
    "province": ("province", "prov"),
    "territoire": ("territoire", "territory", "terr"),
    "commune": ("commune", "municipalite", "municipalité"),
    "quartier": ("quartier", "q", "quartier_avenue"),
    "localite": ("localite", "localité", "locality", "village"),
    "ville": ("ville", "city", "cite", "cité"),
    "site": ("site", "nom_site", "nomsite", "name", "libelle", "libellé", "adresse"),
    "numero": ("numero", "numéro", "n", "id", "code", "code_site"),
}

NOMINATIM_URL = "https://nominatim.openstreetmap.org/search"
NOMINATIM_USER_AGENT = "SIG-FDSU-RDC-Geocoding/1.0 (fdsu-rdc; contact=fdsu@local)"

# Centroïdes de secours pour grandes villes RDC (si PostGIS indisponible)
CITY_FALLBACK: dict[str, tuple[float, float]] = {
    "kinshasa": (-4.325, 15.322),
    "lubumbashi": (-11.664, 27.479),
    "mbuji-mayi": (-6.136, 23.589),
    "kananga": (-5.896, 22.418),
    "kisangani": (0.515, 25.191),
    "bukavu": (-2.491, 28.843),
    "goma": (-1.679, 29.223),
    "matadi": (-5.816, 13.459),
    "mbandaka": (0.049, 18.260),
    "kikwit": (-5.041, 18.816),
    "kindu": (-2.944, 25.922),
    "kolwezi": (-10.715, 25.473),
    "boma": (-5.851, 13.053),
    "moanda": (-5.928, 12.352),
    "ilebo": (-4.331, 20.586),
    "idiofa": (-4.960, 19.590),
    "kasangulu": (-4.583, 15.183),
    "bagata": (-3.315, 17.435),
    "bulungu": (-4.550, 18.600),
    "likasi": (-10.983, 26.733),
    "tshikapa": (-6.416, 20.800),
    "bandundu": (-3.317, 17.375),
    "kalemie": (-5.947, 29.195),
    "butembo": (0.142, 29.291),
    "beni": (0.491, 29.473),
}

STATUS_KEEP = "conservee"
STATUS_CORRECTED = "corrigee"
STATUS_APPROX = "approximative"
STATUS_FAILED = "non_geocodee"
STATUS_SKIPPED = "non_concernee"

_JOBS: dict[str, dict[str, Any]] = {}


@dataclass
class QualityIssue:
    code: str
    message: str
    severity: str = "warning"


@dataclass
class RowGeocodeResult:
    row_number: int
    site_label: str
    old_latitude: Any
    old_longitude: Any
    new_latitude: Any
    new_longitude: Any
    status: str
    source: str | None = None
    confidence_level: str | None = None
    comment: str | None = None
    issues: list[QualityIssue] = field(default_factory=list)
    needs_geocode: bool = False
    address_query: str | None = None
    modified: bool = False

    def to_dict(self) -> dict[str, Any]:
        payload = asdict(self)
        payload["issues"] = [asdict(i) for i in self.issues]
        return payload


def _now() -> str:
    return datetime.now(timezone.utc).isoformat()


def _normalize_header(value: Any) -> str:
    text = str(value or "").strip().lower()
    text = text.replace("é", "e").replace("è", "e").replace("ê", "e").replace("à", "a")
    text = text.replace("ô", "o").replace("û", "u").replace("ç", "c")
    text = re.sub(r"[^a-z0-9]+", "_", text)
    return text.strip("_")


def _to_float(value: Any) -> float | None:
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip().replace(",", ".")
    text = re.sub(r"[^0-9.\-]+", "", text)
    if not text or text in {"-", ".", "-."}:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def _in_rdc(lat: float, lon: float) -> bool:
    return RDC_LAT_MIN <= lat <= RDC_LAT_MAX and RDC_LON_MIN <= lon <= RDC_LON_MAX


def _looks_swapped(lat: float, lon: float) -> bool:
    return (not _in_rdc(lat, lon)) and _in_rdc(lon, lat)


def _is_null_island(lat: float, lon: float) -> bool:
    return abs(lat) < 1e-6 and abs(lon) < 1e-6


def _is_malformed_magnitude(lat: float | None, lon: float | None) -> bool:
    if lat is not None and abs(lat) > 90:
        return True
    if lon is not None and abs(lon) > 180:
        return True
    return False


def ensure_directories() -> None:
    IMPORT_DIR.mkdir(parents=True, exist_ok=True)
    EXPORT_DIR.mkdir(parents=True, exist_ok=True)
    JOBS_DIR.mkdir(parents=True, exist_ok=True)


def detect_columns(headers: list[Any]) -> dict[str, Any]:
    normalized = [_normalize_header(h) for h in headers]
    mapping: dict[str, int | None] = {"latitude": None, "longitude": None}
    for key, aliases in ADDRESS_ALIASES.items():
        mapping[key] = None
        for idx, header in enumerate(normalized):
            if header in aliases or any(alias == header for alias in aliases):
                mapping[key] = idx
                break
            if any(alias in header for alias in aliases if len(alias) > 2):
                mapping[key] = idx
                break

    for idx, header in enumerate(normalized):
        if mapping["latitude"] is None and (header in LAT_ALIASES or header.startswith("lat")):
            mapping["latitude"] = idx
        if mapping["longitude"] is None and (header in LON_ALIASES or header.startswith("lon") or header.startswith("lng")):
            mapping["longitude"] = idx

    return {
        "headers": headers,
        "normalized_headers": normalized,
        "mapping": mapping,
        "has_coordinates": mapping["latitude"] is not None and mapping["longitude"] is not None,
        "address_fields_found": [k for k, v in mapping.items() if v is not None and k not in {"latitude", "longitude"}],
    }


def build_address_query(row_values: list[Any], mapping: dict[str, int | None]) -> str:
    parts: list[str] = []
    for key in ("site", "adresse", "avenue", "quartier", "commune", "localite", "ville", "territoire", "province"):
        idx = mapping.get(key)
        if idx is None:
            continue
        value = row_values[idx] if idx < len(row_values) else None
        text = str(value).strip() if value not in (None, "") else ""
        if text and text not in parts:
            parts.append(text)
    if not parts:
        return "République démocratique du Congo"
    return ", ".join(parts + ["République démocratique du Congo"])


def _pair_key(lat: float | None, lon: float | None) -> str | None:
    if lat is None or lon is None:
        return None
    return f"{round(lat, 5)}|{round(lon, 5)}"


def analyze_coordinate_quality(
    rows: list[list[Any]],
    mapping: dict[str, int | None],
    *,
    repeat_threshold: int = 5,
) -> tuple[list[RowGeocodeResult], dict[str, Any]]:
    lat_idx = mapping.get("latitude")
    lon_idx = mapping.get("longitude")
    site_idx = mapping.get("site") if mapping.get("site") is not None else mapping.get("adresse")
    numero_idx = mapping.get("numero")

    pair_counts: dict[str, int] = {}
    parsed: list[tuple[int, float | None, float | None, list[Any]]] = []

    for offset, row in enumerate(rows):
        row_number = offset + 2  # Excel 1-based + header
        lat = _to_float(row[lat_idx]) if lat_idx is not None and lat_idx < len(row) else None
        lon = _to_float(row[lon_idx]) if lon_idx is not None and lon_idx < len(row) else None
        parsed.append((row_number, lat, lon, row))
        key = _pair_key(lat, lon)
        if key:
            pair_counts[key] = pair_counts.get(key, 0) + 1

    results: list[RowGeocodeResult] = []
    anomaly_counts: dict[str, int] = {}

    for row_number, lat, lon, row in parsed:
        issues: list[QualityIssue] = []
        site_label = ""
        if site_idx is not None and site_idx < len(row) and row[site_idx] not in (None, ""):
            site_label = str(row[site_idx])
        elif numero_idx is not None and numero_idx < len(row):
            site_label = f"Site {row[numero_idx]}"
        else:
            site_label = f"Ligne {row_number}"

        raw_lat = row[lat_idx] if lat_idx is not None and lat_idx < len(row) else None
        raw_lon = row[lon_idx] if lon_idx is not None and lon_idx < len(row) else None
        needs = False
        comment_bits: list[str] = []

        if lat is None or lon is None:
            issues.append(QualityIssue("empty", "Latitude/Longitude vide ou non numérique", "error"))
            needs = True
        else:
            if _is_null_island(lat, lon):
                issues.append(QualityIssue("null_island", "Coordonnées (0,0) invalides", "error"))
                needs = True
            elif _is_malformed_magnitude(lat, lon):
                issues.append(QualityIssue("malformed", "Coordonnées hors plage géographique (±90/±180)", "error"))
                needs = True
            elif not _in_rdc(lat, lon):
                if _looks_swapped(lat, lon):
                    issues.append(QualityIssue("swapped", "Coordonnées probablement inversées (lat/lon)", "error"))
                else:
                    issues.append(QualityIssue("out_of_rdc", "Coordonnées hors emprise RDC", "error"))
                needs = True
            else:
                key = _pair_key(lat, lon)
                if key and pair_counts.get(key, 0) >= repeat_threshold:
                    issues.append(
                        QualityIssue(
                            "repeated",
                            f"Coordonnées génériques répétées {pair_counts[key]} fois",
                            "warning",
                        )
                    )
                    needs = True

        for issue in issues:
            anomaly_counts[issue.code] = anomaly_counts.get(issue.code, 0) + 1
            comment_bits.append(issue.message)

        results.append(
            RowGeocodeResult(
                row_number=row_number,
                site_label=site_label,
                old_latitude=raw_lat,
                old_longitude=raw_lon,
                new_latitude=raw_lat,
                new_longitude=raw_lon,
                status=STATUS_SKIPPED if not needs else STATUS_FAILED,
                issues=issues,
                needs_geocode=needs,
                address_query=build_address_query(row, mapping),
                comment="; ".join(comment_bits) if comment_bits else "Coordonnées valides — conservées",
            )
        )

    summary = {
        "rows_analyzed": len(results),
        "valid_kept_candidate": sum(1 for r in results if not r.needs_geocode),
        "to_geocode": sum(1 for r in results if r.needs_geocode),
        "anomalies": anomaly_counts,
        "repeated_pairs": {
            key: count for key, count in sorted(pair_counts.items(), key=lambda x: -x[1]) if count >= repeat_threshold
        },
    }
    return results, summary


def _normalize_place(value: str) -> str:
    text = value.strip().lower()
    text = text.replace("é", "e").replace("è", "e").replace("ê", "e")
    text = text.replace("à", "a").replace("ô", "o").replace("ç", "c")
    text = re.sub(r"\s+", " ", text)
    return text


def _lookup_offline_place(query_parts: list[str]) -> tuple[float, float, str, str] | None:
    """Retourne (lat, lon, source, confidence) depuis PostGIS ou fallback villes."""
    candidates = [_normalize_place(p) for p in query_parts if p and str(p).strip()]
    if not candidates:
        return None

    if DATA_MODE == "db":
        try:
            with connect_db() as conn:
                with conn.cursor() as cur:
                    for name in candidates:
                        cur.execute(
                            """
                            SELECT ST_Y(ST_Centroid(geom)) AS lat, ST_X(ST_Centroid(geom)) AS lon
                            FROM provinces
                            WHERE lower(coalesce(nom, name, '')) = %s
                              AND geom IS NOT NULL
                            LIMIT 1
                            """,
                            (name,),
                        )
                        row = cur.fetchone()
                        if row and row[0] is not None:
                            return float(row[0]), float(row[1]), "postgis_provinces", "medium"

                        for table, label in (
                            ("territoires", "postgis_territoires"),
                            ("localites", "postgis_localites"),
                            ("villages", "postgis_villages"),
                        ):
                            try:
                                cur.execute("SAVEPOINT geo_place")
                                cur.execute(
                                    f"""
                                    SELECT ST_Y(ST_Centroid(geom)) AS lat, ST_X(ST_Centroid(geom)) AS lon
                                    FROM {table}
                                    WHERE lower(coalesce(nom, name, '')) = %s
                                      AND geom IS NOT NULL
                                    LIMIT 1
                                    """,
                                    (name,),
                                )
                                row = cur.fetchone()
                                cur.execute("RELEASE SAVEPOINT geo_place")
                                if row and row[0] is not None:
                                    return float(row[0]), float(row[1]), label, "medium"
                            except Exception:
                                try:
                                    cur.execute("ROLLBACK TO SAVEPOINT geo_place")
                                except Exception:
                                    pass
        except Exception:
            pass

    for name in candidates:
        if name in CITY_FALLBACK:
            lat, lon = CITY_FALLBACK[name]
            return lat, lon, "offline_city_fallback", "low"
        # fuzzy contains
        for city, coords in CITY_FALLBACK.items():
            if city in name or name in city:
                return coords[0], coords[1], "offline_city_fallback", "low"
    return None


def geocode_nominatim(query: str, *, timeout: float = 20.0) -> tuple[float, float, str, str] | None:
    params = urllib.parse.urlencode(
        {
            "q": query,
            "format": "json",
            "limit": 1,
            "countrycodes": "cd",
        }
    )
    request = urllib.request.Request(
        f"{NOMINATIM_URL}?{params}",
        headers={"User-Agent": NOMINATIM_USER_AGENT, "Accept": "application/json"},
    )
    try:
        with urllib.request.urlopen(request, timeout=timeout) as response:
            payload = json.loads(response.read().decode("utf-8"))
    except Exception:
        return None
    if not payload:
        return None
    hit = payload[0]
    try:
        lat = float(hit["lat"])
        lon = float(hit["lon"])
    except (KeyError, TypeError, ValueError):
        return None
    if not _in_rdc(lat, lon):
        return None
    return lat, lon, "nominatim_osm", "medium"


def _extract_place_parts(row: list[Any], mapping: dict[str, int | None]) -> list[str]:
    parts: list[str] = []
    for key in ("ville", "localite", "commune", "territoire", "quartier", "adresse", "province"):
        idx = mapping.get(key)
        if idx is None or idx >= len(row):
            continue
        value = row[idx]
        if value not in (None, ""):
            parts.append(str(value))
    return parts


def geocode_rows(
    rows: list[list[Any]],
    mapping: dict[str, int | None],
    analysis_results: list[RowGeocodeResult],
    *,
    enable_nominatim: bool = True,
    enable_offline: bool = True,
    max_external_calls: int = 50,
    nominatim_delay_sec: float = 1.05,
) -> list[RowGeocodeResult]:
    external_calls = 0
    output: list[RowGeocodeResult] = []

    for result, row in zip(analysis_results, rows):
        if not result.needs_geocode:
            result.status = STATUS_KEEP
            result.source = "original"
            result.confidence_level = "high"
            result.comment = "Coordonnées valides conservées"
            result.modified = False
            output.append(result)
            continue

        lat = _to_float(result.old_latitude)
        lon = _to_float(result.old_longitude)

        # Correction locale : inversion lat/lon
        if lat is not None and lon is not None and _looks_swapped(lat, lon):
            result.new_latitude = lon
            result.new_longitude = lat
            result.status = STATUS_CORRECTED
            result.source = "swap_lat_lon"
            result.confidence_level = "high"
            result.comment = "Coordonnées inversées corrigées (lat ↔ lon)"
            result.modified = True
            output.append(result)
            continue

        # Magnitude aberrante type -504734 → tentative /100000
        if lat is not None and abs(lat) > 90:
            scaled = lat / 100000.0
            if lon is not None and _in_rdc(scaled, lon):
                result.new_latitude = round(scaled, 6)
                result.new_longitude = lon
                result.status = STATUS_CORRECTED
                result.source = "scale_fix"
                result.confidence_level = "medium"
                result.comment = "Latitude mal formatée recalée (division 1e5)"
                result.modified = True
                output.append(result)
                continue

        resolved = None
        if enable_offline:
            resolved = _lookup_offline_place(_extract_place_parts(row, mapping))

        if resolved is None and enable_nominatim and external_calls < max_external_calls:
            query = result.address_query or build_address_query(row, mapping)
            resolved = geocode_nominatim(query)
            external_calls += 1
            time.sleep(nominatim_delay_sec)

        if resolved is None:
            result.status = STATUS_FAILED
            result.source = None
            result.confidence_level = None
            result.comment = (result.comment or "") + " — géocodage impossible"
            result.modified = False
            output.append(result)
            continue

        new_lat, new_lon, source, confidence = resolved
        result.new_latitude = round(new_lat, 6)
        result.new_longitude = round(new_lon, 6)
        result.source = source
        result.confidence_level = confidence
        result.modified = True
        if confidence == "low" or source.startswith("offline"):
            result.status = STATUS_APPROX
            result.comment = f"Coordonnée approximative via {source}"
        else:
            result.status = STATUS_CORRECTED
            result.comment = f"Coordonnée corrigée via {source}"
        output.append(result)

    return output


def _summarize_results(results: list[RowGeocodeResult]) -> dict[str, Any]:
    return {
        "rows_analyzed": len(results),
        "valid_kept": sum(1 for r in results if r.status == STATUS_KEEP),
        "corrected": sum(1 for r in results if r.status == STATUS_CORRECTED),
        "approximate": sum(1 for r in results if r.status == STATUS_APPROX),
        "failed": sum(1 for r in results if r.status == STATUS_FAILED),
        "modified_cells": sum(1 for r in results if r.modified),
        "anomalies": _count_issues(results),
    }


def _count_issues(results: list[RowGeocodeResult]) -> dict[str, int]:
    counts: dict[str, int] = {}
    for result in results:
        for issue in result.issues:
            counts[issue.code] = counts.get(issue.code, 0) + 1
    return counts


def read_excel_sheet(path: Path, sheet_name: str | None = None) -> tuple[Any, list[Any], list[list[Any]], str]:
    wb = load_workbook(path)
    ws = wb[sheet_name] if sheet_name else wb[wb.sheetnames[0]]
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    rows: list[list[Any]] = []
    for r in range(2, ws.max_row + 1):
        rows.append([ws.cell(r, c).value for c in range(1, ws.max_column + 1)])
    return wb, headers, rows, ws.title


def export_geocoded_workbook(
    source_path: Path,
    results: list[RowGeocodeResult],
    mapping: dict[str, int | None],
    *,
    job_id: str,
) -> Path:
    ensure_directories()
    wb = load_workbook(source_path)
    ws = wb[wb.sheetnames[0]]
    lat_idx = mapping["latitude"]
    lon_idx = mapping["longitude"]
    if lat_idx is None or lon_idx is None:
        raise ValueError("Colonnes Latitude/Longitude introuvables pour l'export.")

    by_row = {r.row_number: r for r in results}
    for row_number, result in by_row.items():
        if not result.modified:
            continue
        # openpyxl columns are 1-based
        ws.cell(row_number, lat_idx + 1).value = result.new_latitude
        ws.cell(row_number, lon_idx + 1).value = result.new_longitude

    # Feuille rapport
    if "GEOCODING_REPORT" in wb.sheetnames:
        del wb["GEOCODING_REPORT"]
    report = wb.create_sheet("GEOCODING_REPORT")
    headers = [
        "numero_ligne",
        "site",
        "ancienne_latitude",
        "ancienne_longitude",
        "nouvelle_latitude",
        "nouvelle_longitude",
        "statut",
        "source",
        "niveau_confiance",
        "commentaire",
    ]
    for col, header in enumerate(headers, start=1):
        cell = report.cell(1, col, header)
        cell.font = Font(bold=True)

    for offset, result in enumerate(results, start=2):
        report.cell(offset, 1, result.row_number)
        report.cell(offset, 2, result.site_label)
        report.cell(offset, 3, result.old_latitude)
        report.cell(offset, 4, result.old_longitude)
        report.cell(offset, 5, result.new_latitude)
        report.cell(offset, 6, result.new_longitude)
        report.cell(offset, 7, result.status)
        report.cell(offset, 8, result.source)
        report.cell(offset, 9, result.confidence_level)
        report.cell(offset, 10, result.comment)

    out_name = f"geocoded_{job_id}.xlsx"
    out_path = EXPORT_DIR / out_name
    wb.save(out_path)
    return out_path


def create_job(kind: str, meta: dict[str, Any] | None = None) -> dict[str, Any]:
    ensure_directories()
    job_id = uuid.uuid4().hex[:12]
    job = {
        "job_id": job_id,
        "kind": kind,
        "status": "created",
        "created_at": _now(),
        "updated_at": _now(),
        "meta": meta or {},
        "summary": {},
        "results_preview": [],
        "export_path": None,
        "error": None,
    }
    _JOBS[job_id] = job
    _persist_job(job)
    return job


def _persist_job(job: dict[str, Any]) -> None:
    ensure_directories()
    path = JOBS_DIR / f"{job['job_id']}.json"
    serializable = dict(job)
    path.write_text(json.dumps(serializable, ensure_ascii=False, indent=2, default=str), encoding="utf-8")


def get_job(job_id: str) -> dict[str, Any] | None:
    if job_id in _JOBS:
        return _JOBS[job_id]
    path = JOBS_DIR / f"{job_id}.json"
    if not path.exists():
        return None
    job = json.loads(path.read_text(encoding="utf-8"))
    _JOBS[job_id] = job
    return job


def analyze_excel_file(path: Path, *, sheet_name: str | None = None) -> dict[str, Any]:
    ensure_directories()
    job = create_job("analyze", {"source_file": str(path)})
    try:
        _wb, headers, rows, used_sheet = read_excel_sheet(path, sheet_name)
        columns = detect_columns(headers)
        if not columns["has_coordinates"]:
            raise ValueError("Colonnes Latitude/Longitude introuvables dans le fichier Excel.")
        results, summary = analyze_coordinate_quality(rows, columns["mapping"])
        suspect = [r.to_dict() for r in results if r.needs_geocode][:100]
        job.update(
            {
                "status": "completed",
                "updated_at": _now(),
                "summary": {
                    **summary,
                    "valid_kept": summary["valid_kept_candidate"],
                    "corrected": 0,
                    "approximate": 0,
                    "failed": summary["to_geocode"],
                    "sheet": used_sheet,
                    "columns": columns,
                },
                "results_preview": suspect,
                "meta": {
                    **job["meta"],
                    "sheet": used_sheet,
                    "columns": columns,
                    "all_results": [r.to_dict() for r in results],
                },
            }
        )
        _persist_job(job)
        return job
    except Exception as exc:
        job["status"] = "failed"
        job["error"] = str(exc)
        job["updated_at"] = _now()
        _persist_job(job)
        raise


def geocode_excel_file(
    path: Path,
    *,
    sheet_name: str | None = None,
    enable_nominatim: bool = False,
    enable_offline: bool = True,
    max_external_calls: int = 30,
) -> dict[str, Any]:
    ensure_directories()
    job = create_job(
        "geocode",
        {
            "source_file": str(path),
            "enable_nominatim": enable_nominatim,
            "enable_offline": enable_offline,
            "max_external_calls": max_external_calls,
        },
    )
    try:
        _wb, headers, rows, used_sheet = read_excel_sheet(path, sheet_name)
        columns = detect_columns(headers)
        if not columns["has_coordinates"]:
            raise ValueError("Colonnes Latitude/Longitude introuvables dans le fichier Excel.")

        analysis_results, _analysis_summary = analyze_coordinate_quality(rows, columns["mapping"])
        results = geocode_rows(
            rows,
            columns["mapping"],
            analysis_results,
            enable_nominatim=enable_nominatim,
            enable_offline=enable_offline,
            max_external_calls=max_external_calls,
        )
        export_path = export_geocoded_workbook(path, results, columns["mapping"], job_id=job["job_id"])
        summary = _summarize_results(results)
        summary.update(
            {
                "sheet": used_sheet,
                "columns": columns,
                "export_path": str(export_path),
                "export_filename": export_path.name,
            }
        )
        preview = [r.to_dict() for r in results if r.modified or r.needs_geocode][:150]
        geojson_features = []
        for result in results:
            lat = _to_float(result.new_latitude)
            lon = _to_float(result.new_longitude)
            if lat is None or lon is None or not _in_rdc(lat, lon):
                continue
            geojson_features.append(
                {
                    "type": "Feature",
                    "geometry": {"type": "Point", "coordinates": [lon, lat]},
                    "properties": {
                        "row_number": result.row_number,
                        "site": result.site_label,
                        "status": result.status,
                        "source": result.source,
                        "confidence_level": result.confidence_level,
                        "modified": result.modified,
                    },
                }
            )

        job.update(
            {
                "status": "completed",
                "updated_at": _now(),
                "summary": summary,
                "results_preview": preview,
                "export_path": str(export_path),
                "geojson": {"type": "FeatureCollection", "features": geojson_features},
                "meta": {
                    **job["meta"],
                    "sheet": used_sheet,
                    "columns": columns,
                    "postgis_ready_rows": [
                        {
                            "site_id": None,
                            "nom_site": r.site_label,
                            "adresse_originale": r.address_query,
                            "latitude": r.new_latitude,
                            "longitude": r.new_longitude,
                            "source_geocoding": r.source,
                            "confidence_level": r.confidence_level,
                            "validation_status": r.status,
                        }
                        for r in results
                        if r.status in {STATUS_KEEP, STATUS_CORRECTED, STATUS_APPROX}
                    ][:500],
                },
            }
        )
        _persist_job(job)
        return job
    except Exception as exc:
        job["status"] = "failed"
        job["error"] = str(exc)
        job["updated_at"] = _now()
        _persist_job(job)
        raise


def resolve_upload_path(filename: str, content: bytes) -> Path:
    ensure_directories()
    safe = re.sub(r"[^A-Za-z0-9._-]+", "_", filename or "upload.xlsx")
    path = IMPORT_DIR / f"{uuid.uuid4().hex[:8]}_{safe}"
    path.write_bytes(content)
    return path


def get_postgis_schema_sql() -> str:
    schema_path = PROJECT_ROOT / "sql" / "geocoding_results.sql"
    if schema_path.exists():
        return schema_path.read_text(encoding="utf-8")
    example_path = PROJECT_ROOT / "docs" / "geocoding_postgis_schema.sql.example"
    if example_path.exists():
        return example_path.read_text(encoding="utf-8")
    return """
CREATE SCHEMA IF NOT EXISTS geocoding;
CREATE TABLE IF NOT EXISTS geocoding.geocoding_results (
    id BIGSERIAL PRIMARY KEY,
    site_id BIGINT NULL,
    nom_site TEXT NULL,
    adresse_originale TEXT NULL,
    latitude DOUBLE PRECISION NULL,
    longitude DOUBLE PRECISION NULL,
    geom geometry(Point, 4326) NULL,
    source_geocoding TEXT NULL,
    confidence_level TEXT NULL,
    validation_status TEXT NULL,
    created_at TIMESTAMPTZ NOT NULL DEFAULT NOW(),
    updated_at TIMESTAMPTZ NOT NULL DEFAULT NOW()
);
""".strip()
