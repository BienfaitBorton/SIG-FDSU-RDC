#!/usr/bin/env python3
"""Pipeline reproductible — Référentiel National des Besoins Numériques (NCI).

Conserve les Excel officiels dans data/raw/.
Produit data/coverage/ (localités, agrégats, qualité, manifeste).
"""

from __future__ import annotations

import hashlib
import json
import math
import re
import sys
from collections import Counter, defaultdict
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

PROJECT_ROOT = Path(__file__).resolve().parents[1]
RAW_DIR = PROJECT_ROOT / "data" / "raw"
OUT_DIR = PROJECT_ROOT / "data" / "coverage"
CONFIG_PATH = OUT_DIR / "nci_config.json"

sys.path.insert(0, str(PROJECT_ROOT))


def _now() -> str:
    return datetime.now(timezone.utc).isoformat()


def _load_config() -> dict[str, Any]:
    return json.loads(CONFIG_PATH.read_text(encoding="utf-8"))


def _find_source(pattern: str) -> Path:
    matches = sorted(RAW_DIR.glob(pattern))
    if not matches:
        raise FileNotFoundError(f"Source introuvable: {RAW_DIR / pattern}")
    return matches[0]


def _sha256(path: Path) -> str:
    h = hashlib.sha256()
    with path.open("rb") as f:
        for chunk in iter(lambda: f.read(1024 * 1024), b""):
            h.update(chunk)
    return h.hexdigest()


def _safe_float(value: Any) -> float | None:
    if value is None or value == "":
        return None
    try:
        v = float(value)
        if math.isnan(v) or math.isinf(v):
            return None
        return v
    except (TypeError, ValueError):
        return None


def _safe_int_pop(value: Any) -> int | None:
    v = _safe_float(value)
    if v is None:
        return None
    return int(round(v))


def _norm_text(value: Any) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def _norm_priority(value: Any, valid: set[str]) -> tuple[str, bool]:
    raw = _norm_text(value)
    if not raw:
        return "unknown", False
    # Exact / case-insensitive match
    for v in valid:
        if raw.lower() == v.lower():
            return v, True
    return "invalid", False


def _norm_category(value: Any) -> str:
    raw = _norm_text(value)
    if not raw:
        return "unknown"
    if ">10000" in raw.replace(" ", ""):
        return ">10000"
    m = re.search(r"([A-Ea-e])\s*$", raw) or re.search(r"[Cc]at[ée]gorie\s*([A-Ea-e])", raw)
    if m:
        return m.group(1).upper()
    # Already single letter
    if len(raw) == 1 and raw.upper() in "ABCDE":
        return raw.upper()
    return raw


def _coords_ok(lat: float | None, lon: float | None) -> bool:
    if lat is None or lon is None:
        return False
    # DRC approximate bounding box (with margin)
    return -14.0 <= lat <= 6.0 and 11.0 <= lon <= 32.5


def _quality_score(row: dict[str, Any], checks: dict[str, int], valid_priorities: set[str]) -> dict[str, Any]:
    parts: dict[str, bool] = {
        "coordinates": bool(row.get("coords_valid")),
        "territory": bool(row.get("territoire")),
        "province": bool(row.get("province")),
        "population": row.get("population") is not None and row.get("population") >= 0,
        "priority": row.get("priority") in valid_priorities,
        "category": row.get("categorie") not in (None, "unknown") if "categorie" in row else True,
        "infrastructure": bool(row.get("infra_name") or row.get("infra_type")),
    }
    score = 0
    detail = {}
    for key, weight in checks.items():
        ok = parts.get(key, False)
        detail[key] = {"ok": ok, "weight": weight}
        if ok:
            score += weight
    return {"score": score, "checks": detail}


def _row_id(dataset: str, idx: int, name: str | None, lat: float | None, lon: float | None) -> str:
    base = f"{dataset}|{idx}|{name or ''}|{lat}|{lon}"
    return f"NCI-{dataset[:3].upper()}-{hashlib.md5(base.encode()).hexdigest()[:10].upper()}"


def _parse_uncovered(path: Path, config: dict[str, Any]) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    import openpyxl

    valid_pri = set(config.get("valid_priorities") or ["High", "Medium", "Low"])
    checks = (config.get("data_quality") or {}).get("checks") or {}
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    # Prefer sheet with localities
    sheet_name = next((s for s in wb.sheetnames if "localit" in s.lower() or "non couvert" in s.lower()), wb.sheetnames[-1])
    ws = wb[sheet_name]
    rows: list[dict[str, Any]] = []
    issues = Counter()
    seen_keys: dict[str, int] = {}

    for i, raw in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not raw or all(c is None or str(c).strip() == "" for c in raw):
            continue
        name = _norm_text(raw[1] if len(raw) > 1 else None)
        lat = _safe_float(raw[2] if len(raw) > 2 else None)
        lon = _safe_float(raw[3] if len(raw) > 3 else None)
        project = _norm_text(raw[4] if len(raw) > 4 else None)
        priority, pri_ok = _norm_priority(raw[5] if len(raw) > 5 else None, valid_pri)
        destination = _norm_text(raw[6] if len(raw) > 6 else None)
        distance = _safe_float(raw[7] if len(raw) > 7 else None)
        infra_name = _norm_text(raw[8] if len(raw) > 8 else None)
        infra_distance = _safe_float(raw[9] if len(raw) > 9 else None)
        infra_type = _norm_text(raw[10] if len(raw) > 10 else None)
        zone = _norm_text(raw[11] if len(raw) > 11 else None)
        province = _norm_text(raw[12] if len(raw) > 12 else None)
        territoire = _norm_text(raw[13] if len(raw) > 13 else None)
        population = _safe_int_pop(raw[14] if len(raw) > 14 else None)
        categorie = _norm_category(raw[15] if len(raw) > 15 else None)

        coords_valid = _coords_ok(lat, lon)
        if not coords_valid:
            issues["invalid_coords"] += 1
        if not pri_ok:
            issues["invalid_priority"] += 1
        if not province:
            issues["missing_province"] += 1
        if not territoire:
            issues["missing_territory"] += 1
        if population is None:
            issues["missing_population"] += 1

        dup_key = f"{(name or '').lower()}|{lat}|{lon}|{territoire or ''}"
        is_dup = dup_key in seen_keys
        if is_dup:
            issues["duplicates"] += 1
        else:
            seen_keys[dup_key] = i

        row = {
            "id": _row_id("uncovered", i, name, lat, lon),
            "dataset": "localities_uncovered",
            "coverage_status": "uncovered",
            "source_row": i,
            "name": name,
            "latitude": lat,
            "longitude": lon,
            "coords_valid": coords_valid,
            "project": project,
            "priority": priority,
            "priority_valid": pri_ok,
            "destination": destination,
            "distance_km": distance,
            "infra_name": infra_name,
            "infra_distance_km": infra_distance,
            "infra_type": infra_type,
            "fdsu_zone": zone,
            "province": province,
            "territoire": territoire,
            "population": population,
            "categorie": categorie,
            "duplicate": is_dup,
        }
        row["data_quality"] = _quality_score(row, checks, valid_pri)
        rows.append(row)

    wb.close()
    report = {
        "sheet": sheet_name,
        "rows": len(rows),
        "issues": dict(issues),
        "unique_keys": len(seen_keys),
    }
    return rows, report


def _parse_population_coverage(path: Path, config: dict[str, Any]) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    import openpyxl

    valid_pri = set(config.get("valid_priorities") or ["High", "Medium", "Low"])
    checks = dict((config.get("data_quality") or {}).get("checks") or {})
    # Population coverage sheet has no category column — treat as not applicable (full points).
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows: list[dict[str, Any]] = []
    issues = Counter()
    seen_keys: dict[str, int] = {}

    for i, raw in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not raw or all(c is None or str(c).strip() == "" for c in raw):
            continue
        # No '#' column — indices start at Site Name
        name = _norm_text(raw[0] if len(raw) > 0 else None)
        lat = _safe_float(raw[1] if len(raw) > 1 else None)
        lon = _safe_float(raw[2] if len(raw) > 2 else None)
        project = _norm_text(raw[3] if len(raw) > 3 else None)
        priority, pri_ok = _norm_priority(raw[4] if len(raw) > 4 else None, valid_pri)
        destination = _norm_text(raw[5] if len(raw) > 5 else None)
        distance = _safe_float(raw[6] if len(raw) > 6 else None)
        infra_name = _norm_text(raw[7] if len(raw) > 7 else None)
        infra_distance = _safe_float(raw[8] if len(raw) > 8 else None)
        infra_type = _norm_text(raw[9] if len(raw) > 9 else None)
        zone = _norm_text(raw[10] if len(raw) > 10 else None)
        province = _norm_text(raw[11] if len(raw) > 11 else None)
        territoire = _norm_text(raw[12] if len(raw) > 12 else None)
        population = _safe_int_pop(raw[13] if len(raw) > 13 else None)

        coords_valid = _coords_ok(lat, lon)
        if not coords_valid:
            issues["invalid_coords"] += 1
        if not pri_ok:
            issues["invalid_priority"] += 1
        if not province:
            issues["missing_province"] += 1
        if not territoire:
            issues["missing_territory"] += 1
        if population is None:
            issues["missing_population"] += 1

        dup_key = f"{(name or '').lower()}|{lat}|{lon}|{territoire or ''}"
        is_dup = dup_key in seen_keys
        if is_dup:
            issues["duplicates"] += 1
        else:
            seen_keys[dup_key] = i

        row = {
            "id": _row_id("covered", i, name, lat, lon),
            "dataset": "population_coverage",
            "coverage_status": "covered",
            "source_row": i,
            "name": name,
            "latitude": lat,
            "longitude": lon,
            "coords_valid": coords_valid,
            "project": project,
            "priority": priority,
            "priority_valid": pri_ok,
            "destination": destination,
            "distance_km": distance,
            "infra_name": infra_name,
            "infra_distance_km": infra_distance,
            "infra_type": infra_type,
            "fdsu_zone": zone,
            "province": province,
            "territoire": territoire,
            "population": population,
            "categorie": "N/A",
            "duplicate": is_dup,
        }
        dq = _quality_score(row, checks, valid_pri)
        if "category" in checks:
            dq["checks"]["category"] = {
                "ok": True,
                "weight": checks["category"],
                "applicable": False,
            }
            # Ensure category weight is counted when not applicable
            if not dq["checks"].get("category", {}).get("ok"):
                pass
            base = sum(
                c["weight"]
                for k, c in dq["checks"].items()
                if k != "category" and c.get("ok")
            )
            dq["score"] = base + checks["category"]
        row["data_quality"] = dq
        rows.append(row)

    wb.close()
    report = {
        "sheet": ws.title,
        "rows": len(rows),
        "issues": dict(issues),
        "unique_keys": len(seen_keys),
    }
    return rows, report


def _aggregate(rows: list[dict[str, Any]], status: str) -> dict[str, Any]:
    by_province: dict[str, dict[str, Any]] = defaultdict(lambda: {
        "localities": 0, "population": 0, "priorities": Counter(), "categories": Counter(),
        "distances": [], "infra_types": Counter(),
    })
    by_territory: dict[str, dict[str, Any]] = defaultdict(lambda: {
        "localities": 0, "population": 0, "priorities": Counter(), "categories": Counter(),
        "distances": [], "infra_types": Counter(), "province": None, "fdsu_zone": None,
    })
    priorities: Counter = Counter()
    categories: Counter = Counter()
    infra_types: Counter = Counter()
    distances: list[float] = []
    quality_scores: list[int] = []
    pop_total = 0
    coords_ok = 0

    for r in rows:
        if r.get("duplicate"):
            continue
        pop = r.get("population") or 0
        pop_total += pop
        pri = r.get("priority") or "unknown"
        priorities[pri] += 1
        cat = r.get("categorie")
        if cat:
            categories[cat] += 1
        it = r.get("infra_type") or "unknown"
        infra_types[it] += 1
        if r.get("distance_km") is not None:
            distances.append(float(r["distance_km"]))
        if r.get("coords_valid"):
            coords_ok += 1
        dq = (r.get("data_quality") or {}).get("score")
        if isinstance(dq, (int, float)):
            quality_scores.append(int(dq))

        prov = r.get("province") or "INCONNU"
        terr = r.get("territoire") or "INCONNU"
        bp = by_province[prov]
        bp["localities"] += 1
        bp["population"] += pop
        bp["priorities"][pri] += 1
        if cat:
            bp["categories"][cat] += 1
        if r.get("distance_km") is not None:
            bp["distances"].append(float(r["distance_km"]))
        bp["infra_types"][it] += 1

        bt = by_territory[terr]
        bt["localities"] += 1
        bt["population"] += pop
        bt["priorities"][pri] += 1
        if cat:
            bt["categories"][cat] += 1
        if r.get("distance_km") is not None:
            bt["distances"].append(float(r["distance_km"]))
        bt["infra_types"][it] += 1
        bt["province"] = r.get("province")
        bt["fdsu_zone"] = r.get("fdsu_zone")

    def _finalize_geo(d: dict[str, Any]) -> dict[str, Any]:
        dists = d.pop("distances")
        return {
            "localities": d["localities"],
            "population": d["population"],
            "priorities": dict(d["priorities"]),
            "categories": dict(d["categories"]),
            "infra_types": dict(d["infra_types"]),
            "avg_distance_km": round(sum(dists) / len(dists), 2) if dists else None,
            "province": d.get("province"),
            "fdsu_zone": d.get("fdsu_zone"),
        }

    return {
        "coverage_status": status,
        "localities": len([r for r in rows if not r.get("duplicate")]),
        "localities_including_duplicates": len(rows),
        "population": pop_total,
        "coords_valid": coords_ok,
        "avg_distance_km": round(sum(distances) / len(distances), 2) if distances else None,
        "avg_data_quality": round(sum(quality_scores) / len(quality_scores), 1) if quality_scores else None,
        "priorities": dict(priorities),
        "categories": dict(categories),
        "infra_types": dict(infra_types.most_common(50)),
        "by_province": {k: _finalize_geo(v) for k, v in sorted(by_province.items())},
        "by_territory": {k: _finalize_geo(v) for k, v in sorted(by_territory.items())},
    }


def _compute_ndci(uncovered_terr: dict[str, Any], covered_terr: dict[str, Any] | None, config: dict[str, Any]) -> dict[str, Any]:
    idx = config["national_digital_coverage_index"]
    weights = idx["weights"]
    pri_scores = idx["priority_scores"]
    cat_scores = idx["category_scores"]
    dist_cfg = idx["distance_bands_km"]
    infra_scores = idx["infrastructure_scores"]

    unc = uncovered_terr or {}
    cov = covered_terr or {}
    pop_unc = int(unc.get("population") or 0)
    pop_cov = int(cov.get("population") or 0)
    pop_tot = pop_unc + pop_cov
    # Higher remaining need → higher index (needs intensity)
    pop_score = min(100.0, (pop_unc / max(pop_tot, 1)) * 100.0) if pop_tot else 0.0

    pri = unc.get("priorities") or {}
    pri_weighted = 0.0
    pri_n = 0
    for k, n in pri.items():
        pri_weighted += pri_scores.get(k, pri_scores.get("unknown", 10)) * n
        pri_n += n
    priority_score = (pri_weighted / pri_n) if pri_n else 0.0

    cats = unc.get("categories") or {}
    cat_weighted = 0.0
    cat_n = 0
    for k, n in cats.items():
        cat_weighted += cat_scores.get(k, cat_scores.get("unknown", 10)) * n
        cat_n += n
    category_score = (cat_weighted / cat_n) if cat_n else 0.0

    avg_d = unc.get("avg_distance_km")
    if avg_d is None:
        distance_score = dist_cfg["scores"]["missing"]
    elif avg_d <= dist_cfg["near_max"]:
        distance_score = dist_cfg["scores"]["near"]
    elif avg_d <= dist_cfg["medium_max"]:
        distance_score = dist_cfg["scores"]["medium"]
    else:
        distance_score = dist_cfg["scores"]["far"]

    infra = unc.get("infra_types") or {}
    known = sum(v for k, v in infra.items() if k and k != "unknown")
    total_infra = sum(infra.values()) or 0
    if total_infra == 0:
        infrastructure_score = infra_scores["unknown"]
    elif known / total_infra >= 0.5:
        infrastructure_score = infra_scores["present"]
    else:
        infrastructure_score = infra_scores["absent"]

    wsum = sum(weights.values()) or 100
    index = (
        weights["population"] * pop_score
        + weights["priority"] * priority_score
        + weights["category"] * category_score
        + weights["distance"] * distance_score
        + weights["infrastructure"] * infrastructure_score
    ) / wsum

    return {
        "index": round(index, 1),
        "components": {
            "population": round(pop_score, 1),
            "priority": round(priority_score, 1),
            "category": round(category_score, 1),
            "distance": round(distance_score, 1),
            "infrastructure": round(infrastructure_score, 1),
        },
        "weights": weights,
        "version": idx.get("version"),
        "id": idx.get("id"),
        "population_uncovered": pop_unc,
        "population_covered": pop_cov,
        "localities_uncovered": unc.get("localities") or 0,
        "localities_covered": cov.get("localities") or 0,
    }


def _write_jsonl(path: Path, rows: list[dict[str, Any]]) -> None:
    with path.open("w", encoding="utf-8") as f:
        for row in rows:
            f.write(json.dumps(row, ensure_ascii=False) + "\n")


def main() -> int:
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    config = _load_config()

    uncovered_src = _find_source("Localit*.xlsx")
    coverage_src = _find_source("Population*.xlsx")

    print(f"Import uncovered: {uncovered_src.name}")
    uncovered, unc_report = _parse_uncovered(uncovered_src, config)
    print(f"  -> {len(uncovered)} rows")

    print(f"Import population coverage: {coverage_src.name}")
    covered, cov_report = _parse_population_coverage(coverage_src, config)
    print(f"  -> {len(covered)} rows")

    unc_agg = _aggregate(uncovered, "uncovered")
    cov_agg = _aggregate(covered, "covered")

    # NDCI per territory
    all_terr_names = set(unc_agg["by_territory"]) | set(cov_agg["by_territory"])
    ndci_by_territory = {}
    for name in sorted(all_terr_names):
        ndci_by_territory[name] = _compute_ndci(
            unc_agg["by_territory"].get(name),
            cov_agg["by_territory"].get(name),
            config,
        )

    # Merge province rollups for national view
    provinces = sorted(set(unc_agg["by_province"]) | set(cov_agg["by_province"]))
    by_province_merged = {}
    for p in provinces:
        u = unc_agg["by_province"].get(p) or {}
        c = cov_agg["by_province"].get(p) or {}
        by_province_merged[p] = {
            "province": p,
            "localities_uncovered": u.get("localities") or 0,
            "localities_covered": c.get("localities") or 0,
            "population_uncovered": u.get("population") or 0,
            "population_covered": c.get("population") or 0,
            "population_remaining": u.get("population") or 0,
            "priorities_uncovered": u.get("priorities") or {},
            "categories_uncovered": u.get("categories") or {},
            "avg_distance_km_uncovered": u.get("avg_distance_km"),
            "avg_distance_km_covered": c.get("avg_distance_km"),
        }

    territories_merged = {}
    for name in sorted(all_terr_names):
        u = unc_agg["by_territory"].get(name) or {}
        c = cov_agg["by_territory"].get(name) or {}
        ndci = ndci_by_territory[name]
        territories_merged[name] = {
            "territoire": name,
            "province": u.get("province") or c.get("province"),
            "fdsu_zone": u.get("fdsu_zone") or c.get("fdsu_zone"),
            "localities_uncovered": u.get("localities") or 0,
            "localities_covered": c.get("localities") or 0,
            "population_uncovered": u.get("population") or 0,
            "population_covered": c.get("population") or 0,
            "population_remaining": u.get("population") or 0,
            "priorities": u.get("priorities") or {},
            "categories": u.get("categories") or {},
            "avg_distance_km": u.get("avg_distance_km"),
            "infra_types": u.get("infra_types") or {},
            "ndci": ndci,
            "data_quality_avg": None,
        }

    # Average quality per territory from uncovered rows
    q_by_terr: dict[str, list[int]] = defaultdict(list)
    for r in uncovered:
        if r.get("duplicate"):
            continue
        t = r.get("territoire") or "INCONNU"
        q = (r.get("data_quality") or {}).get("score")
        if isinstance(q, (int, float)):
            q_by_terr[t].append(int(q))
    for t, scores in q_by_terr.items():
        if t in territories_merged:
            territories_merged[t]["data_quality_avg"] = round(sum(scores) / len(scores), 1)

    national = {
        "localities_uncovered": unc_agg["localities"],
        "localities_covered": cov_agg["localities"],
        "population_uncovered": unc_agg["population"],
        "population_covered": cov_agg["population"],
        "population_remaining": unc_agg["population"],
        "population_total_observed": unc_agg["population"] + cov_agg["population"],
        "avg_distance_km_uncovered": unc_agg["avg_distance_km"],
        "avg_distance_km_covered": cov_agg["avg_distance_km"],
        "avg_data_quality_uncovered": unc_agg["avg_data_quality"],
        "avg_data_quality_covered": cov_agg["avg_data_quality"],
        "priorities_uncovered": unc_agg["priorities"],
        "priorities_covered": cov_agg["priorities"],
        "categories_uncovered": unc_agg["categories"],
        "infra_types_uncovered": unc_agg["infra_types"],
        "coverage_ratio_localities": round(
            cov_agg["localities"] / max(cov_agg["localities"] + unc_agg["localities"], 1), 4
        ),
        "coverage_ratio_population": round(
            cov_agg["population"] / max(cov_agg["population"] + unc_agg["population"], 1), 4
        ),
    }

    aggregates = {
        "_meta": {
            "generated_at": _now(),
            "engine": "import_national_coverage",
            "heritage": "Référentiel National des Besoins",
        },
        "national": national,
        "by_province": by_province_merged,
        "by_territory": territories_merged,
        "uncovered": {k: v for k, v in unc_agg.items() if k not in {"by_province", "by_territory"}},
        "covered": {k: v for k, v in cov_agg.items() if k not in {"by_province", "by_territory"}},
        "ndci_top_territories": sorted(
            (
                {"territoire": k, "province": territories_merged[k].get("province"), **v}
                for k, v in ndci_by_territory.items()
            ),
            key=lambda x: x.get("index") or 0,
            reverse=True,
        )[:30],
    }

    quality_report = {
        "_meta": {"generated_at": _now(), "id": "CDQS_V1"},
        "uncovered": unc_report,
        "covered": cov_report,
        "national_avg_quality": {
            "uncovered": unc_agg["avg_data_quality"],
            "covered": cov_agg["avg_data_quality"],
        },
    }

    manifest = {
        "_meta": {
            "title": "National Coverage Intelligence — manifeste",
            "generated_at": _now(),
            "schema_version": "1.0.0",
        },
        "heritage": "Référentiel National des Besoins Numériques",
        "sources": [
            {
                "role": "localities_uncovered",
                "path": str(uncovered_src.relative_to(PROJECT_ROOT)).replace("\\", "/"),
                "sha256": _sha256(uncovered_src),
                "preserved_original": True,
            },
            {
                "role": "population_coverage",
                "path": str(coverage_src.relative_to(PROJECT_ROOT)).replace("\\", "/"),
                "sha256": _sha256(coverage_src),
                "preserved_original": True,
            },
        ],
        "outputs": {
            "localities_uncovered": "data/coverage/localities_uncovered.jsonl",
            "localities_covered": "data/coverage/localities_covered.jsonl",
            "aggregates": "data/coverage/aggregates.json",
            "quality_report": "data/coverage/quality_report.json",
            "config": "data/coverage/nci_config.json",
        },
        "counts": {
            "localities_uncovered": national["localities_uncovered"],
            "localities_covered": national["localities_covered"],
            "population_uncovered": national["population_uncovered"],
            "population_covered": national["population_covered"],
        },
    }

    _write_jsonl(OUT_DIR / "localities_uncovered.jsonl", uncovered)
    _write_jsonl(OUT_DIR / "localities_covered.jsonl", covered)
    (OUT_DIR / "aggregates.json").write_text(
        json.dumps(aggregates, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    (OUT_DIR / "quality_report.json").write_text(
        json.dumps(quality_report, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    (OUT_DIR / "manifest.json").write_text(
        json.dumps(manifest, ensure_ascii=False, indent=2), encoding="utf-8"
    )

    print("OK - NCI import termine")
    print(json.dumps(manifest["counts"], ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
