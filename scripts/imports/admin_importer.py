from __future__ import annotations

import argparse
from collections import defaultdict
from pathlib import Path

from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.exc import SQLAlchemyError
from sqlalchemy.orm import Session
from tqdm import tqdm

from app.database import engine
from app.models import (
    Collectivite,
    CollectiviteType,
    Groupement,
    Province,
    Territoire,
    Village,
)


class ImportReport:
    def __init__(self) -> None:
        self.inserted = defaultdict(int)
        self.skipped = defaultdict(int)
        self.errors: list[str] = []

    def add_inserted(self, model_name: str) -> None:
        self.inserted[model_name] += 1

    def add_skipped(self, model_name: str) -> None:
        self.skipped[model_name] += 1

    def add_error(self, message: str) -> None:
        self.errors.append(message)

    def summary(self) -> str:
        lines = ["Rapport d'import administratif FDSU RDC :"]
        for model_name in ["Province", "Territoire", "Collectivite", "Groupement", "Village"]:
            lines.append(
                f"- {model_name} : insérés = {self.inserted[model_name]}, ignorés = {self.skipped[model_name]}"
            )
        if self.errors:
            lines.append(f"- erreurs = {len(self.errors)}")
            lines.extend(f"  - {error}" for error in self.errors)
        else:
            lines.append("- erreurs = 0")
        return "\n".join(lines)


def normalize_text(value: object | None) -> str:
    if value is None:
        return ""
    return str(value).strip()


def get_cell_value(row: tuple, column_index: int) -> str:
    return normalize_text(row[column_index].value)


def import_provinces(workbook, report: ImportReport, session: Session) -> None:
    worksheet = workbook["Provinces"]
    rows = list(worksheet.iter_rows(min_row=2, values_only=False))
    for row in tqdm(rows, desc="Import Provinces", unit="province"):
        nom = normalize_text(row[0].value)
        code = normalize_text(row[1].value)
        zone = normalize_text(row[2].value)
        chef_lieu = normalize_text(row[3].value) or None
        population = normalize_text(row[4].value)
        superficie = normalize_text(row[5].value)

        if not nom or not code or not zone:
            report.add_skipped("Province")
            continue

        existing = session.scalar(select(Province).where(Province.code == code))
        if existing:
            report.add_skipped("Province")
            continue

        try:
            province = Province(
                nom=nom,
                code=code,
                zone=zone,
                chef_lieu=chef_lieu,
                population=int(population) if population else None,
                superficie=float(superficie) if superficie else None,
            )
            session.add(province)
            report.add_inserted("Province")
        except Exception as exc:  # pragma: no cover
            report.add_error(f"Province {code} : {exc}")


def import_territoires(workbook, report: ImportReport, session: Session) -> None:
    worksheet = workbook["Territoires"]
    rows = list(worksheet.iter_rows(min_row=2, values_only=False))
    for row in tqdm(rows, desc="Import Territoires", unit="territoire"):
        nom = normalize_text(row[0].value)
        code = normalize_text(row[1].value)
        province_code = normalize_text(row[2].value)
        chef_lieu = normalize_text(row[3].value) or None

        if not nom or not code or not province_code:
            report.add_skipped("Territoire")
            continue

        province = session.scalar(select(Province).where(Province.code == province_code))
        if province is None:
            report.add_error(
                f"Territoire {code} ignoré : province '{province_code}' introuvable"
            )
            continue

        existing = session.scalar(select(Territoire).where(Territoire.code == code, Territoire.province_id == province.id))
        if existing:
            report.add_skipped("Territoire")
            continue

        try:
            territoire = Territoire(
                nom=nom,
                code=code,
                chef_lieu=chef_lieu,
                province_id=province.id,
            )
            session.add(territoire)
            report.add_inserted("Territoire")
        except Exception as exc:  # pragma: no cover
            report.add_error(f"Territoire {code} : {exc}")


def import_collectivites(workbook, report: ImportReport, session: Session) -> None:
    worksheet = workbook["Collectivites"]
    rows = list(worksheet.iter_rows(min_row=2, values_only=False))
    for row in tqdm(rows, desc="Import Collectivites", unit="collectivite"):
        nom = normalize_text(row[0].value)
        code = normalize_text(row[1].value)
        type_collectivite = normalize_text(row[2].value)
        territoire_code = normalize_text(row[3].value)

        if not nom or not code or not type_collectivite or not territoire_code:
            report.add_skipped("Collectivite")
            continue

        try:
            collectivite_type = CollectiviteType(type_collectivite)
        except ValueError:
            report.add_error(
                f"Collectivite {code} ignorée : type invalide '{type_collectivite}'"
            )
            continue

        territoire = session.scalar(select(Territoire).where(Territoire.code == territoire_code))
        if territoire is None:
            report.add_error(
                f"Collectivite {code} ignorée : territoire '{territoire_code}' introuvable"
            )
            continue

        existing = session.scalar(
            select(Collectivite).where(
                Collectivite.code == code,
                Collectivite.territoire_id == territoire.id,
            )
        )
        if existing:
            report.add_skipped("Collectivite")
            continue

        try:
            collectivite = Collectivite(
                nom=nom,
                code=code,
                type_collectivite=collectivite_type,
                territoire_id=territoire.id,
            )
            session.add(collectivite)
            report.add_inserted("Collectivite")
        except Exception as exc:  # pragma: no cover
            report.add_error(f"Collectivite {code} : {exc}")


def import_groupements(workbook, report: ImportReport, session: Session) -> None:
    worksheet = workbook["Groupements"]
    rows = list(worksheet.iter_rows(min_row=2, values_only=False))
    for row in tqdm(rows, desc="Import Groupements", unit="groupement"):
        nom = normalize_text(row[0].value)
        code = normalize_text(row[1].value)
        collectivite_code = normalize_text(row[2].value)

        if not nom or not code or not collectivite_code:
            report.add_skipped("Groupement")
            continue

        collectivite = session.scalar(select(Collectivite).where(Collectivite.code == collectivite_code))
        if collectivite is None:
            report.add_error(
                f"Groupement {code} ignoré : collectivite '{collectivite_code}' introuvable"
            )
            continue

        existing = session.scalar(
            select(Groupement).where(
                Groupement.code == code,
                Groupement.collectivite_id == collectivite.id,
            )
        )
        if existing:
            report.add_skipped("Groupement")
            continue

        try:
            groupement = Groupement(
                nom=nom,
                code=code,
                collectivite_id=collectivite.id,
            )
            session.add(groupement)
            report.add_inserted("Groupement")
        except Exception as exc:  # pragma: no cover
            report.add_error(f"Groupement {code} : {exc}")


def import_villages(workbook, report: ImportReport, session: Session) -> None:
    worksheet = workbook["Villages"]
    rows = list(worksheet.iter_rows(min_row=2, values_only=False))
    for row in tqdm(rows, desc="Import Villages", unit="village"):
        nom = normalize_text(row[0].value)
        code = normalize_text(row[1].value)
        groupement_code = normalize_text(row[2].value)

        if not nom or not code or not groupement_code:
            report.add_skipped("Village")
            continue

        groupement = session.scalar(select(Groupement).where(Groupement.code == groupement_code))
        if groupement is None:
            report.add_error(
                f"Village {code} ignoré : groupement '{groupement_code}' introuvable"
            )
            continue

        existing = session.scalar(
            select(Village).where(
                Village.code == code,
                Village.groupement_id == groupement.id,
            )
        )
        if existing:
            report.add_skipped("Village")
            continue

        try:
            village = Village(
                nom=nom,
                code=code,
                groupement_id=groupement.id,
            )
            session.add(village)
            report.add_inserted("Village")
        except Exception as exc:  # pragma: no cover
            report.add_error(f"Village {code} : {exc}")


def import_administratif(file_path: Path) -> ImportReport:
    report = ImportReport()
    workbook = load_workbook(filename=file_path, data_only=True)

    with Session(engine) as session:
        try:
            with session.begin():
                import_provinces(workbook, report, session)
                import_territoires(workbook, report, session)
                import_collectivites(workbook, report, session)
                import_groupements(workbook, report, session)
                import_villages(workbook, report, session)
        except SQLAlchemyError as exc:
            session.rollback()
            report.add_error(f"Échec transactionnel : {exc}")

    return report


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Importer les référentiels administratifs RDC depuis un fichier Excel." 
    )
    parser.add_argument(
        "file",
        type=Path,
        help="Chemin vers le fichier .xlsx contenant les référentiels administratifs.",
    )
    args = parser.parse_args()

    file_path = args.file
    if not file_path.exists():
        raise FileNotFoundError(f"Fichier introuvable : {file_path}")

    report = import_administratif(file_path)
    print(report.summary())


if __name__ == "__main__":
    main()
