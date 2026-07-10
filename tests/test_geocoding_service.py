"""Tests unitaires — Géocodage Intelligent FDSU."""

from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook

from api.services import geocoding_service as geo


@pytest.fixture()
def sample_workbook(tmp_path: Path) -> Path:
    path = tmp_path / "sites_test.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Export"
    ws.append(["Numero", "Adresse", "Latitude", "Longitude", "Province", "Quartier", "Ville"])
    ws.append([1, "Site valide Kinshasa", -4.325, 15.322, "Kinshasa", "Gombe", "Kinshasa"])
    ws.append([2, "Site zero", 0, 0, "Equateur", "Centre", "Mbandaka"])
    ws.append([3, "Site hors RDC", 48.85, 2.35, "Paris", "x", "Paris"])
    ws.append([4, "Latitude mal formée", -504734, 18.77735, "Bandundu", "LUKOLELA", "Kikwit"])
    ws.append([5, "Répété A", -3.31576, 17.43497, "Bandundu", "A", "Bagata"])
    ws.append([6, "Répété B", -3.31576, 17.43497, "Bandundu", "B", "Bulungu"])
    ws.append([7, "Répété C", -3.31576, 17.43497, "Bandundu", "C", "Idiofa"])
    ws.append([8, "Répété D", -3.31576, 17.43497, "Bandundu", "D", "Kikwit"])
    ws.append([9, "Répété E", -3.31576, 17.43497, "Bandundu", "E", "Kikwit"])
    ws.append([10, "Vide", None, None, "Haut-Katanga", "x", "Lubumbashi"])
    # Colonne métier à ne pas modifier
    ws["H1"] = "CodeMetier"
    for row in range(2, 12):
        ws.cell(row, 8, f"KEEP-{row}")
    wb.save(path)
    return path


def test_detect_invalid_and_keep_valid(sample_workbook: Path):
    _wb, headers, rows, _sheet = geo.read_excel_sheet(sample_workbook)
    columns = geo.detect_columns(headers)
    assert columns["has_coordinates"] is True
    results, summary = geo.analyze_coordinate_quality(rows, columns["mapping"], repeat_threshold=5)

    valid = [r for r in results if not r.needs_geocode]
    assert len(valid) == 1
    assert valid[0].old_latitude == -4.325

    codes = {issue.code for r in results for issue in r.issues}
    assert "null_island" in codes
    assert "out_of_rdc" in codes
    assert "malformed" in codes
    assert "empty" in codes
    assert "repeated" in codes
    assert summary["to_geocode"] >= 5


def test_geocode_preserves_valid_and_exports_report(sample_workbook: Path, tmp_path: Path, monkeypatch):
    # Forcer offline city fallback sans réseau
    monkeypatch.setattr(geo, "DATA_MODE", "json")

    job = geo.geocode_excel_file(
        sample_workbook,
        enable_nominatim=False,
        enable_offline=True,
        max_external_calls=0,
    )
    assert job["status"] == "completed"
    summary = job["summary"]
    assert summary["valid_kept"] >= 1
    assert summary["corrected"] + summary["approximate"] >= 1
    assert Path(job["export_path"]).exists()

    wb = load_workbook(job["export_path"])
    assert "GEOCODING_REPORT" in wb.sheetnames
    ws = wb["Export"]
    # Ligne valide conservée
    assert float(ws.cell(2, 3).value) == pytest.approx(-4.325)
    assert float(ws.cell(2, 4).value) == pytest.approx(15.322)
    # Autres colonnes non modifiées
    assert ws.cell(2, 8).value == "KEEP-2"
    assert ws.cell(3, 8).value == "KEEP-3"
    # Ligne (0,0) Mbandaka approximée/corrigée
    assert ws.cell(3, 3).value not in (0, 0.0, None)
    assert ws.cell(3, 4).value not in (0, 0.0, None)

    report = wb["GEOCODING_REPORT"]
    assert report.cell(1, 1).value == "numero_ligne"
    assert report.max_row >= 2


def test_other_columns_untouched_on_analyze_only(sample_workbook: Path):
    job = geo.analyze_excel_file(sample_workbook)
    assert job["status"] == "completed"
    assert job["summary"]["rows_analyzed"] == 10
    # analyze ne produit pas d'export écrasant
    assert job.get("export_path") in (None, "")
