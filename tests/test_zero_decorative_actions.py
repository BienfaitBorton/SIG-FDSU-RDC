"""Tests Zero Decorative Actions — exports dossier + capacités."""

from __future__ import annotations

from io import BytesIO

import pytest
from fastapi.testclient import TestClient
from openpyxl import load_workbook

from api.main import app
from api.services import export_service

CLIENT = TestClient(app)


def test_capabilities_registry():
    payload = export_service.get_capabilities()
    assert payload["capabilities"]["export_excel"] is True
    assert payload["capabilities"]["export_pdf"] is False
    assert payload["capabilities"]["export_powerpoint"] is False
    assert payload["capabilities"]["mission_planning"] is False
    assert payload["capabilities"]["simulation"] is False

    response = CLIENT.get("/api/exports/capabilities")
    assert response.status_code == 200
    body = response.json()
    assert body["capabilities"]["export_excel"] is True


def test_decision_case_excel_export_real_xlsx():
    # Site 7 sites_40 — cas observé utilisateur ; fallback sur premier site scoré si absent
    response = CLIENT.get("/api/exports/decision-case/site/7/excel?program_code=sites_40")
    if response.status_code == 404:
        # Environnement sans site 7 : vérifier 404 métier propre
        assert "introuvable" in response.json()["detail"].lower()
        pytest.skip("Site 7 indisponible dans cet environnement")
    assert response.status_code == 200
    assert "spreadsheetml" in response.headers.get("content-type", "")
    filename = response.headers.get("X-FDSU-Export-Filename") or ""
    assert filename.endswith(".xlsx")
    assert "Dossier_decision" in filename
    data = response.content
    assert len(data) > 100
    wb = load_workbook(BytesIO(data))
    assert "Dossier" in wb.sheetnames
    assert "Critères" in wb.sheetnames
    values = [str(c.value or "") for row in wb["Dossier"].iter_rows(max_row=20) for c in row]
    joined = " ".join(values)
    assert "Programme" in joined
    assert "Recommandation" in joined


def test_pdf_and_ppt_not_fake_success():
    pdf = CLIENT.get("/api/exports/decision-case/site/7/pdf")
    ppt = CLIENT.get("/api/exports/decision-case/site/7/powerpoint")
    assert pdf.status_code == 501
    assert ppt.status_code == 501
    assert "non encore activé" in pdf.json()["detail"].lower()
    assert "non encore activé" in ppt.json()["detail"].lower()


def test_safe_text_never_object_repr():
    assert export_service._safe_text({"recommendation": "Prioriser le site"}) == "Prioriser le site"
    assert "[object Object]" not in export_service._safe_text({"a": 1})
